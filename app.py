
import json
import os
import random
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from flask import Flask, abort, flash, g, jsonify, redirect, render_template, request, session, url_for

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
DATA_DIR = BASE_DIR / "data"
SEED_FILE = Path(os.environ.get("SEED_XLSX_PATH", str(BASE_DIR / "sample_data" / "Career_Crox_Interns_Seed.xlsx")))
DB_PATH = DATA_DIR / "career_crox_demo.db"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")
USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)

app = Flask(__name__, template_folder=str(BASE_DIR / "templates"), static_folder=str(BASE_DIR / "static"))
app.secret_key = os.environ.get("SECRET_KEY", "career-crox-demo-secret-key")

SIDEBAR_ITEMS = [
    ("Dashboard", "dashboard", {}),
    ("Candidates", "candidates", {}),
    ("JD Centre", "jds", {}),
    ("Interviews", "interviews", {}),
    ("Tasks", "tasks", {}),
    ("Dialer", "module_page", {"slug": "dialer"}),
    ("Meeting Room", "module_page", {"slug": "meeting-room"}),
    ("Learning Hub", "module_page", {"slug": "learning-hub"}),
    ("Social Career Crox", "module_page", {"slug": "social-career-crox"}),
    ("Wallet & Rewards", "module_page", {"slug": "wallet-rewards"}),
    ("Payout Tracker", "module_page", {"slug": "payout-tracker"}),
    ("Reports", "module_page", {"slug": "reports"}),
    ("Admin Control", "admin_page", {}),
    ("Mega Blueprint", "blueprint_page", {}),
]

MODULE_SUMMARIES = {
    "dialer": {
        "title": "Dialer Command Center",
        "summary": "Call outcomes, talktime, callback suggestions, and recruiter productivity in one place.",
        "cards": [("Connected Today", "64"), ("Callbacks Due", "18"), ("Talktime", "06h 42m"), ("Best Hour", "10-11 AM")],
        "items": ["One-click dial flow", "Call outcome popup", "Hourly scoreboard", "Talktime analytics", "Follow-up suggestion engine"]
    },
    "meeting-room": {
        "title": "Meeting Room",
        "summary": "Create, join, and monitor internal meetings with attendance and quick notes.",
        "cards": [("Meetings Today", "7"), ("Live Rooms", "2"), ("Attendance Rate", "91%"), ("Pending Minutes", "3")],
        "items": ["Create Meeting", "Join Meeting", "Attendance", "Raise Hand", "Meeting chat"]
    },
    "learning-hub": {
        "title": "Learning Hub",
        "summary": "Training videos, process notes, and coaching clips for recruiters and TLs.",
        "cards": [("Videos", "24"), ("Playlists", "6"), ("Completion", "72%"), ("New This Week", "4")],
        "items": ["Airtel process videos", "Interview objection handling", "Salary negotiation tips", "Manager coaching clips"]
    },
    "social-career-crox": {
        "title": "Social Career Crox",
        "summary": "Plan social posts, manage queue, and track posting status across platforms.",
        "cards": [("Queued Posts", "12"), ("Posted", "41"), ("Missed", "2"), ("This Week Reach", "8.2k")],
        "items": ["Schedule Post", "Post Queue", "Posted Posts", "Missed Posts", "Platform filters"]
    },
    "wallet-rewards": {
        "title": "Wallet & Rewards",
        "summary": "Trips, reward milestones, and streak-based motivation without the fake motivational posters.",
        "cards": [("Reward Budget", "₹42,000"), ("Eligible Recruiters", "6"), ("Nearest Milestone", "2 joinings"), ("Trips Active", "3")],
        "items": ["20 Joining → Goa Trip", "10 Interview Conversions → Bonus", "Monthly target streak", "Reward history"]
    },
    "payout-tracker": {
        "title": "Payout Tracker",
        "summary": "Eligibility, confirmations, invoice readiness, and team-wise payout visibility.",
        "cards": [("Eligible Profiles", "11"), ("Client Confirmations", "8"), ("Invoice Ready", "6"), ("Pending Cases", "5")],
        "items": ["Recruiter earning view", "Team earnings", "Target vs achieved", "60-day eligible tracker", "Dispute notes"]
    },
    "reports": {
        "title": "Reports",
        "summary": "Funnel, conversion, source, and location reports for managers who enjoy charts more than chaos.",
        "cards": [("Lead → Join", "8.6%"), ("Top Source", "Naukri"), ("Top City", "Noida"), ("Top Recruiter", "Ritika")],
        "items": ["Daily report", "Weekly funnel", "Source performance", "Location analytics", "Recruiter efficiency"]
    }
}

SAMPLE_PUBLIC_NOTES = [
    ("C001", "recruiter.01", "public", "Candidate confirmed she can attend interview tomorrow around 11 AM.", -2),
    ("C003", "recruiter.02", "public", "Strong communication and ready for final round.", -1),
    ("C004", "admin", "public", "Reschedule approved. Keep candidate warm and confirm next slot.", -1),
]

SAMPLE_PRIVATE_NOTES = [
    ("C001", "recruiter.01", "private", "Responds faster after 6 PM. Daytime follow-up is messy.", -1),
    ("C003", "admin", "private", "Useful benchmark profile for this process.", -2),
]

SAMPLE_MESSAGES = [
    ("admin", "tl.noida", "Please review pending Airtel profiles today.", -1),
    ("tl.noida", "recruiter.01", "Do a proper callback and update note history.", -1),
    ("recruiter.01", "tl.noida", "Done. Candidate is responsive.", 0),
]

THEME_ALIASES = {
    "midnight": "dark-midnight",
    "cobalt": "ocean",
    "graphite": "dark-pro",
    "forest": "mint",
    "sunset": "sunset",
    "lavender": "lavender",
    "silver": "silver-pro",
    "silver-pro": "silver-pro",
    "dark": "dark-pro",
    "dark-pro": "dark-pro",
    "dark-midnight": "dark-midnight",
    "dark-vscode": "dark-vscode",
    "ocean": "ocean",
    "mint": "mint",
    "rose": "rose",
    "corporate-light": "corporate-light",
}

ALLOWED_THEMES = {
    "corporate-light", "ocean", "rose", "mint", "sunset",
    "lavender", "dark-pro", "dark-midnight", "dark-vscode", "silver-pro"
}

TABLE_COLUMNS = {
    "users": {"user_id","username","password","full_name","designation","role","recruiter_code","is_active","theme_name","updated_at"},
    "candidates": {"candidate_id","full_name","phone","qualification","location","experience","preferred_location","qualification_level","total_experience","relevant_experience","in_hand_salary","career_gap","documents_availability","communication_skill","relevant_experience_range","relevant_in_hand_range","submission_date","process","recruiter_code","recruiter_name","recruiter_designation","status","all_details_sent","interview_reschedule_date","is_duplicate","notes","resume_filename","recording_filename","created_at","updated_at"},
    "tasks": {"task_id","title","description","assigned_to_user_id","assigned_to_name","assigned_by_user_id","assigned_by_name","status","priority","due_date","created_at","updated_at"},
    "notifications": {"notification_id","user_id","title","message","category","status","metadata","created_at"},
    "jd_master": {"jd_id","job_title","company","location","experience","salary","notes","created_at"},
    "settings": {"setting_key","setting_value","notes","Instructions"},
    "notes": {"candidate_id","username","note_type","body","created_at"},
    "messages": {"sender_username","recipient_username","body","created_at"},
    "interviews": {"interview_id","candidate_id","jd_id","stage","scheduled_at","status","created_at"},
    "submissions": {"submission_id","candidate_id","jd_id","recruiter_code","status","submitted_at"},
}

def trim_to_columns(rows, table_name):
    allowed = TABLE_COLUMNS[table_name]
    out = []
    for row in rows:
        out.append({k: row.get(k, "") for k in allowed if k in row})
    return out


def now_iso():
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

def display_ts(value, default=""):
    if not value:
        return default
    text = str(value).replace("T", " ").strip()
    return text[:16]

def normalize_theme(theme_name):
    theme_name = (theme_name or "").strip()
    mapped = THEME_ALIASES.get(theme_name, theme_name)
    return mapped if mapped in ALLOWED_THEMES else "corporate-light"

def normalize_role(role):
    role = (role or "").strip().lower()
    if role in {"admin", "manager", "operations", "ops"}:
        return "manager"
    if role in {"tl", "team lead", "lead"}:
        return "tl"
    return "recruiter"

def to_boolish(val):
    return str(val).strip().lower() in {"1", "true", "yes", "y"}

def clean_row(row):
    cleaned = {}
    for k, v in row.items():
        if k is None:
            continue
        if isinstance(v, datetime):
            cleaned[k] = v.isoformat()
        elif v is None:
            cleaned[k] = ""
        else:
            cleaned[k] = str(v).strip() if not isinstance(v, (int, float)) else str(v)
    return cleaned

def parse_sheet_rows(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h is not None else None for h in raw_headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        item = {}
        for idx, value in enumerate(row):
            key = headers[idx] if idx < len(headers) else None
            if not key:
                continue
            item[key] = value
        if item:
            rows.append(clean_row(item))
    wb.close()
    return rows

def derive_seed_payload(xlsx_path):
    users = parse_sheet_rows(xlsx_path, "Users")
    candidates = parse_sheet_rows(xlsx_path, "Candidates")
    tasks = parse_sheet_rows(xlsx_path, "Tasks")
    notifications = parse_sheet_rows(xlsx_path, "Notifications")
    jds = parse_sheet_rows(xlsx_path, "JD_Master")
    settings = parse_sheet_rows(xlsx_path, "Settings")

    if not users:
        users = [
            {"user_id": "U001", "username": "admin", "password": "Admin@123", "full_name": "Ava Operations", "designation": "Admin", "role": "admin", "recruiter_code": "ADMIN", "is_active": "1", "theme_name": "corporate-light", "updated_at": now_iso()},
            {"user_id": "U002", "username": "tl.noida", "password": "TL@12345", "full_name": "Rhea Team Lead", "designation": "Team Lead", "role": "tl", "recruiter_code": "TL-ND", "is_active": "1", "theme_name": "corporate-light", "updated_at": now_iso()},
            {"user_id": "U003", "username": "recruiter.01", "password": "Rec@12345", "full_name": "Arjun Recruiter", "designation": "Recruiter", "role": "recruiter", "recruiter_code": "RC-101", "is_active": "1", "theme_name": "corporate-light", "updated_at": now_iso()},
        ]
    if not candidates:
        candidates = [
            {"candidate_id": "C001", "full_name": "Neha Sharma", "phone": "9876543210", "qualification": "B.A.", "location": "Delhi", "experience": "6 Months", "preferred_location": "Noida", "qualification_level": "Graduate", "total_experience": "6", "relevant_experience": "4", "in_hand_salary": "18000", "career_gap": "Currently Working", "documents_availability": "Yes", "communication_skill": "Good", "relevant_experience_range": "4-6 Months", "relevant_in_hand_range": "16-20K", "submission_date": datetime.now().strftime("%Y-%m-%d"), "process": "Airtel", "recruiter_code": "RC-101", "recruiter_name": "Arjun Recruiter", "recruiter_designation": "Recruiter", "status": "New", "all_details_sent": "Pending", "interview_reschedule_date": "", "is_duplicate": "0", "notes": "Interested in voice process", "resume_filename": "", "recording_filename": "", "created_at": now_iso(), "updated_at": now_iso()},
        ]
    if not tasks:
        tasks = [
            {"task_id": "T001", "title": "Verify pending Airtel profiles", "description": "Review document status for Airtel candidates.", "assigned_to_user_id": "U003", "assigned_to_name": "Arjun Recruiter", "assigned_by_user_id": "U002", "assigned_by_name": "Rhea Team Lead", "status": "Open", "priority": "High", "due_date": datetime.now().strftime("%Y-%m-%d"), "created_at": now_iso(), "updated_at": now_iso()}
        ]
    if not notifications:
        notifications = [
            {"notification_id": "N001", "user_id": "U003", "title": "Seed imported", "message": "Candidate and user data loaded into CRM.", "category": "system", "status": "Unread", "metadata": "{}", "created_at": now_iso()}
        ]
    if not jds:
        jds = [
            {"jd_id": "J001", "job_title": "Customer Support Associate", "company": "Airtel", "location": "Noida", "experience": "0-12 Months", "salary": "16K-22K", "notes": "Voice support process", "created_at": now_iso()}
        ]
    if not settings:
        settings = [
            {"setting_key": "company_name", "setting_value": "Career Crox Interns", "notes": "Replace with your company name", "Instructions": "Supabase-backed CRM demo"},
            {"setting_key": "default_theme", "setting_value": "corporate-light", "notes": "One of the built-in themes", "Instructions": "User-specific themes are stored in users.theme_name"},
        ]

    usernames = [u.get("username", "") for u in users]
    candidate_codes = [c.get("candidate_id", "") for c in candidates]
    now = datetime.now()
    notes = []
    for candidate_id, username, note_type, body, day_offset in SAMPLE_PUBLIC_NOTES + SAMPLE_PRIVATE_NOTES:
        if candidate_id in candidate_codes and username in usernames:
            notes.append({
                "candidate_id": candidate_id,
                "username": username,
                "note_type": note_type,
                "body": body,
                "created_at": (now + timedelta(days=day_offset, hours=random.randint(8, 18))).isoformat(timespec="seconds")
            })

    messages = []
    for sender, recipient, body, day_offset in SAMPLE_MESSAGES:
        if sender in usernames and recipient in usernames:
            messages.append({
                "sender_username": sender,
                "recipient_username": recipient,
                "body": body,
                "created_at": (now + timedelta(days=day_offset, hours=random.randint(8, 18))).isoformat(timespec="seconds")
            })

    interviews = []
    submissions = []
    for idx, c in enumerate(candidates, start=1):
        candidate_id = c.get("candidate_id") or f"C{idx:03d}"
        process = c.get("process") or ""
        jd_match = next((j for j in jds if (j.get("company") or "").strip().lower() == process.strip().lower()), None)
        jd_id = jd_match.get("jd_id") if jd_match else ""
        status = c.get("status") or "New"
        submitted_at = c.get("submission_date") or now.strftime("%Y-%m-%d")
        submissions.append({
            "submission_id": f"S{idx:03d}",
            "candidate_id": candidate_id,
            "jd_id": jd_id,
            "recruiter_code": c.get("recruiter_code", ""),
            "status": status,
            "submitted_at": submitted_at
        })
        if "interview" in status.lower() or c.get("interview_reschedule_date"):
            when = c.get("interview_reschedule_date") or f"{submitted_at} 11:00"
            interviews.append({
                "interview_id": f"I{idx:03d}",
                "candidate_id": candidate_id,
                "jd_id": jd_id,
                "stage": status,
                "scheduled_at": when,
                "status": "Scheduled" if "reschedule" not in status.lower() else "Rescheduled",
                "created_at": now_iso()
            })

    return {
        "users": trim_to_columns(users, "users"),
        "candidates": trim_to_columns(candidates, "candidates"),
        "tasks": trim_to_columns(tasks, "tasks"),
        "notifications": trim_to_columns(notifications, "notifications"),
        "jd_master": trim_to_columns(jds, "jd_master"),
        "settings": trim_to_columns(settings, "settings"),
        "notes": trim_to_columns(notes, "notes"),
        "messages": trim_to_columns(messages, "messages"),
        "interviews": trim_to_columns(interviews, "interviews"),
        "submissions": trim_to_columns(submissions, "submissions"),
    }

class SQLiteBackend:
    def __init__(self, db_path, seed_file):
        self.db_path = str(db_path)
        self.seed_file = Path(seed_file)
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self._init_db()
        self._seed_if_empty()

    def describe(self):
        return {"store_mode": "sqlite-demo", "seed_file": str(self.seed_file)}

    def _connect(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        conn = self._connect()
        cur = conn.cursor()
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            full_name TEXT,
            designation TEXT,
            role TEXT,
            recruiter_code TEXT,
            is_active TEXT,
            theme_name TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS candidates (
            candidate_id TEXT PRIMARY KEY,
            full_name TEXT,
            phone TEXT,
            qualification TEXT,
            location TEXT,
            experience TEXT,
            preferred_location TEXT,
            qualification_level TEXT,
            total_experience TEXT,
            relevant_experience TEXT,
            in_hand_salary TEXT,
            career_gap TEXT,
            documents_availability TEXT,
            communication_skill TEXT,
            relevant_experience_range TEXT,
            relevant_in_hand_range TEXT,
            submission_date TEXT,
            process TEXT,
            recruiter_code TEXT,
            recruiter_name TEXT,
            recruiter_designation TEXT,
            status TEXT,
            all_details_sent TEXT,
            interview_reschedule_date TEXT,
            is_duplicate TEXT,
            notes TEXT,
            resume_filename TEXT,
            recording_filename TEXT,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS tasks (
            task_id TEXT PRIMARY KEY,
            title TEXT,
            description TEXT,
            assigned_to_user_id TEXT,
            assigned_to_name TEXT,
            assigned_by_user_id TEXT,
            assigned_by_name TEXT,
            status TEXT,
            priority TEXT,
            due_date TEXT,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS notifications (
            notification_id TEXT PRIMARY KEY,
            user_id TEXT,
            title TEXT,
            message TEXT,
            category TEXT,
            status TEXT,
            metadata TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS jd_master (
            jd_id TEXT PRIMARY KEY,
            job_title TEXT,
            company TEXT,
            location TEXT,
            experience TEXT,
            salary TEXT,
            notes TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS settings (
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT,
            notes TEXT,
            Instructions TEXT
        );
        CREATE TABLE IF NOT EXISTS notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id TEXT,
            username TEXT,
            note_type TEXT,
            body TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_username TEXT,
            recipient_username TEXT,
            body TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS interviews (
            interview_id TEXT PRIMARY KEY,
            candidate_id TEXT,
            jd_id TEXT,
            stage TEXT,
            scheduled_at TEXT,
            status TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS submissions (
            submission_id TEXT PRIMARY KEY,
            candidate_id TEXT,
            jd_id TEXT,
            recruiter_code TEXT,
            status TEXT,
            submitted_at TEXT
        );
        """)
        conn.commit()
        conn.close()

    def _seed_if_empty(self):
        if self.count("users") > 0:
            return
        payload = derive_seed_payload(self.seed_file) if self.seed_file.exists() else derive_seed_payload("")
        for table, rows in payload.items():
            if rows:
                self.bulk_insert(table, rows)

    def count(self, table):
        conn = self._connect()
        cur = conn.execute(f"SELECT COUNT(*) as c FROM {table}")
        c = cur.fetchone()["c"]
        conn.close()
        return c

    def list_rows(self, table):
        conn = self._connect()
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM {table}").fetchall()]
        conn.close()
        return rows

    def bulk_insert(self, table, rows):
        if not rows:
            return
        conn = self._connect()
        keys = list(rows[0].keys())
        placeholders = ",".join(["?"] * len(keys))
        sql = f"INSERT INTO {table} ({','.join(keys)}) VALUES ({placeholders})"
        vals = [[row.get(k, "") for k in keys] for row in rows]
        conn.executemany(sql, vals)
        conn.commit()
        conn.close()

    def insert(self, table, row):
        keys = list(row.keys())
        placeholders = ",".join(["?"] * len(keys))
        conn = self._connect()
        conn.execute(f"INSERT INTO {table} ({','.join(keys)}) VALUES ({placeholders})", [row.get(k, "") for k in keys])
        conn.commit()
        conn.close()

    def update_where(self, table, filters, values):
        if not values:
            return
        set_sql = ", ".join([f"{k}=?" for k in values.keys()])
        where_sql = " AND ".join([f"{k}=?" for k in filters.keys()])
        params = list(values.values()) + list(filters.values())
        conn = self._connect()
        conn.execute(f"UPDATE {table} SET {set_sql} WHERE {where_sql}", params)
        conn.commit()
        conn.close()

class SupabaseBackend:
    def __init__(self, url, service_role_key):
        self.url = url.rstrip("/")
        self.key = service_role_key
        self.headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

    def describe(self):
        return {"store_mode": "supabase", "url": self.url}

    def _request(self, method, path, params=None, json_body=None):
        resp = requests.request(
            method,
            f"{self.url}/rest/v1/{path}",
            headers=self.headers,
            params=params or {},
            json=json_body,
            timeout=20,
        )
        if resp.status_code >= 400:
            raise RuntimeError(f"Supabase {method} {path} failed: {resp.status_code} {resp.text[:250]}")
        if not resp.text:
            return []
        try:
            return resp.json()
        except Exception:
            return []

    def list_rows(self, table):
        return self._request("GET", table, params={"select": "*"})

    def insert(self, table, row):
        self._request("POST", table, json_body=row)

    def bulk_insert(self, table, rows):
        if rows:
            self._request("POST", table, json_body=rows)

    def update_where(self, table, filters, values):
        params = {"select": "*"}
        for k, v in filters.items():
            params[k] = f"eq.{v}"
        self._request("PATCH", table, params=params, json_body=values)

def get_backend():
    if "backend" not in g:
        if USE_SUPABASE:
            g.backend = SupabaseBackend(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
        else:
            g.backend = SQLiteBackend(DB_PATH, SEED_FILE)
    return g.backend

@app.teardown_appcontext
def close_backend(error=None):
    g.pop("backend", None)

def list_users():
    rows = get_backend().list_rows("users")
    normalized = []
    for row in rows:
        item = dict(row)
        item["role"] = normalize_role(item.get("role"))
        item["app_role"] = item["role"]
        item["theme_name"] = normalize_theme(item.get("theme_name"))
        item["is_active"] = "1" if to_boolish(item.get("is_active", "1")) else "0"
        normalized.append(item)
    return normalized

def user_map(by="username"):
    rows = list_users()
    return {u.get(by): u for u in rows if u.get(by)}

def get_user(username):
    return user_map("username").get(username)

def find_user_by_recruiter_code(code):
    code = (code or "").strip()
    for user in list_users():
        if (user.get("recruiter_code") or "").strip() == code:
            return user
    return None

def visible_private_notes(candidate_id, user):
    notes = [n for n in get_backend().list_rows("notes") if (n.get("candidate_id") or "") == candidate_id and (n.get("note_type") or "") == "private"]
    users = user_map("username")
    if user["role"] != "manager":
        notes = [n for n in notes if (n.get("username") or "") == user["username"]]
    notes.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    for n in notes:
        u = users.get(n.get("username")) or {}
        n["full_name"] = u.get("full_name", n.get("username"))
        n["designation"] = u.get("designation", "")
        n["created_at"] = display_ts(n.get("created_at"))
    return notes

def recruiters_for_filters():
    items = []
    for u in list_users():
        if u["role"] == "recruiter":
            items.append({"username": u.get("recruiter_code") or u.get("username"), "full_name": u.get("full_name", "")})
    return items

def enrich_candidates():
    users_by_code = {u.get("recruiter_code"): u for u in list_users() if u.get("recruiter_code")}
    tl_users = [u for u in list_users() if u["role"] == "tl"]
    jds = get_backend().list_rows("jd_master")
    rows = get_backend().list_rows("candidates")
    enriched = []
    for row in rows:
        item = dict(row)
        item["code"] = item.get("candidate_id", "")
        item["jd_code"] = item.get("process", "")
        item["created_at"] = display_ts(item.get("created_at"))
        item["updated_at"] = display_ts(item.get("updated_at"))
        item["recruiter_code"] = item.get("recruiter_code", "")
        item["experience"] = item.get("experience") or item.get("total_experience") or ""
        user = users_by_code.get(item["recruiter_code"]) or {}
        item["recruiter_name"] = item.get("recruiter_name") or user.get("full_name", "")
        item["recruiter_designation"] = item.get("recruiter_designation") or user.get("designation", "")
        tl = tl_users[0] if tl_users else {}
        item["tl_name"] = tl.get("full_name", "")
        item["tl_username"] = tl.get("username", "")
        jd = next((j for j in jds if (j.get("company") or "").strip().lower() == (item.get("process") or "").strip().lower()), None)
        item["jd_title"] = f"{jd.get('job_title')} • {jd.get('company')}" if jd else (item.get("process") or "")
        item["payout"] = jd.get("salary", "") if jd else ""
        item["jd_status"] = "Open"
        enriched.append(item)
    return enriched

def candidate_map():
    return {c["code"]: c for c in enrich_candidates()}

def get_candidate(code):
    return candidate_map().get(code)

def enrich_notifications():
    users_by_id = user_map("user_id")
    out = []
    for row in get_backend().list_rows("notifications"):
        item = dict(row)
        user = users_by_id.get(item.get("user_id")) or {}
        item["username"] = user.get("username", "")
        item["is_read"] = 1 if (item.get("status") or "").lower() == "read" else 0
        item["created_at"] = display_ts(item.get("created_at"))
        out.append(item)
    out.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return out

def user_notifications(user, candidate_code=None, unread_only=False):
    notifications = enrich_notifications()
    items = []
    for n in notifications:
        if n.get("username") != user["username"]:
            continue
        if candidate_code:
            metadata = n.get("metadata") or ""
            if candidate_code not in metadata and candidate_code != n.get("candidate_id"):
                continue
        if unread_only and n["is_read"]:
            continue
        items.append(n)
    return items

def current_user():
    uname = session.get("impersonated_as") or session.get("username")
    return get_user(uname) if uname else None

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def manager_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user or user["role"] != "manager":
            abort(403)
        return fn(*args, **kwargs)
    return wrapper

@app.context_processor
def inject_globals():
    user = current_user()
    unread = len(user_notifications(user, unread_only=True)) if user else 0
    active_theme = normalize_theme((user or {}).get("theme_name"))
    return {
        "sidebar_items": SIDEBAR_ITEMS,
        "current_user_data": user,
        "unread_notifications": unread,
        "now": datetime.now(),
        "active_theme": active_theme,
    }

@app.route("/health")
def health():
    try:
        info = get_backend().describe()
        return {"ok": True, **info}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}, 500

@app.route("/debug_store")
def debug_store():
    info = get_backend().describe()
    sample_users = list_users()[:5]
    sample_candidates = enrich_candidates()[:5]
    return jsonify({"store": info, "users": sample_users, "candidates": sample_candidates})

@app.route("/api/theme", methods=["POST"])
@login_required
def save_theme():
    user = current_user()
    payload = request.get_json(silent=True) or {}
    theme = normalize_theme(payload.get("theme"))
    if theme not in ALLOWED_THEMES:
        return jsonify({"ok": False}), 400
    get_backend().update_where("users", {"user_id": user["user_id"]}, {"theme_name": theme, "updated_at": now_iso()})
    session["theme_name"] = theme
    return jsonify({"ok": True, "theme": theme})

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = get_user(username)
        if user and user.get("password") == password and to_boolish(user.get("is_active", "1")):
            session.clear()
            session["username"] = user["username"]
            session["theme_name"] = normalize_theme(user.get("theme_name"))
            flash(f"Welcome back, {user['full_name']}. The dashboard was waiting, tragically.", "success")
            return redirect(url_for("dashboard"))
        flash("Invalid login. Check the username/password from your imported Users sheet.", "danger")
    demo_users = sorted(list_users(), key=lambda x: (x["role"], x.get("full_name", "")))
    return render_template("login.html", demo_users=demo_users)

@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out.", "info")
    return redirect(url_for("login"))

@app.route("/")
@login_required
def root():
    return redirect(url_for("dashboard"))

@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    candidates = enrich_candidates()
    interviews = get_backend().list_rows("interviews")
    tasks = get_backend().list_rows("tasks")
    users = list_users()

    total_profiles = len([c for c in candidates if not to_boolish(c.get("is_duplicate", "0"))])
    today_calls = max(12, len(candidates) * 3)
    today_str = datetime.now().strftime("%Y-%m-%d")
    interviews_today = len([i for i in interviews if today_str in str(i.get("scheduled_at", ""))])
    active_managers = len([u for u in users if u["role"] in {"manager", "tl"}])

    recent_activity = sorted(candidates, key=lambda x: x.get("created_at", ""), reverse=True)[:6]
    due_tasks = []
    user_by_id = user_map("user_id")
    for task in tasks:
        t = dict(task)
        assigned = user_by_id.get(t.get("assigned_to_user_id")) or {}
        t["full_name"] = assigned.get("full_name", t.get("assigned_to_name", ""))
        t["assigned_to"] = assigned.get("username", "")
        t["due_at"] = t.get("due_date", "")
        due_tasks.append(t)
    due_tasks.sort(key=lambda x: (x.get("status", ""), x.get("due_at", "")))
    due_tasks = due_tasks[:6]

    manager_monitoring = []
    for u in users:
        if u["role"] not in {"recruiter", "tl"}:
            continue
        ccount = len([c for c in candidates if (c.get("recruiter_code") or "") == (u.get("recruiter_code") or "")])
        open_tasks = len([t for t in tasks if t.get("assigned_to_user_id") == u.get("user_id") and (t.get("status") or "") != "Closed"])
        manager_monitoring.append({"full_name": u.get("full_name"), "designation": u.get("designation"), "candidate_count": ccount, "open_tasks": open_tasks})
    manager_monitoring.sort(key=lambda x: (-x["candidate_count"], x["full_name"]))
    manager_monitoring = manager_monitoring[:6]

    unread_notes = user_notifications(user)[:5]
    return render_template("dashboard.html",
        total_profiles=total_profiles,
        today_calls=today_calls,
        interviews_today=interviews_today,
        active_managers=active_managers,
        recent_activity=recent_activity,
        due_tasks=due_tasks,
        manager_monitoring=manager_monitoring,
        unread_notes=unread_notes
    )

@app.route("/candidates")
@login_required
def candidates():
    q = request.args.get("q", "").strip().lower()
    recruiter = request.args.get("recruiter", "").strip()
    status = request.args.get("status", "").strip()
    rows = [c for c in enrich_candidates() if not to_boolish(c.get("is_duplicate", "0"))]
    if q:
        rows = [c for c in rows if q in " ".join([c.get("full_name",""), c.get("phone",""), c.get("location",""), c.get("status",""), c.get("jd_code",""), c.get("recruiter_code","")]).lower()]
    if recruiter:
        rows = [c for c in rows if c.get("recruiter_code") == recruiter]
    if status:
        rows = [c for c in rows if c.get("status") == status]
    rows.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    statuses = sorted({c.get("status", "") for c in enrich_candidates() if c.get("status")})
    return render_template("candidates.html", candidates=rows, q=request.args.get("q",""), recruiters=recruiters_for_filters(), current_recruiter=recruiter, statuses=statuses, current_status=status)

@app.route("/candidate/<candidate_code>")
@login_required
def candidate_detail(candidate_code):
    user = current_user()
    candidate = get_candidate(candidate_code)
    if not candidate:
        abort(404)
    notes = get_backend().list_rows("notes")
    users = user_map("username")
    public_notes = []
    for n in notes:
        if n.get("candidate_id") != candidate_code or n.get("note_type") != "public":
            continue
        item = dict(n)
        u = users.get(item.get("username")) or {}
        item["full_name"] = u.get("full_name", item.get("username", ""))
        item["designation"] = u.get("designation", "")
        item["created_at"] = display_ts(item.get("created_at"))
        public_notes.append(item)
    public_notes.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    private_notes = visible_private_notes(candidate_code, user)
    related_notifications = user_notifications(user, candidate_code=candidate_code)[:8]

    timeline = []
    for s in get_backend().list_rows("submissions"):
        if s.get("candidate_id") == candidate_code:
            timeline.append({"event_type": "Submission", "label": s.get("status", ""), "event_time": display_ts(s.get("submitted_at")), "jd_code": s.get("jd_id", ""), "owner": s.get("recruiter_code","")})
    for i in get_backend().list_rows("interviews"):
        if i.get("candidate_id") == candidate_code:
            timeline.append({"event_type": "Interview", "label": i.get("status", ""), "event_time": display_ts(i.get("scheduled_at")), "jd_code": i.get("jd_id", ""), "owner": ""})
    timeline.sort(key=lambda x: x.get("event_time", ""), reverse=True)

    return render_template("candidate_detail.html",
                           candidate=candidate,
                           public_notes=public_notes,
                           private_notes=private_notes,
                           related_notifications=related_notifications,
                           timeline=timeline)

@app.route("/candidate/<candidate_code>/add-note", methods=["POST"])
@login_required
def add_note(candidate_code):
    user = current_user()
    note_type = request.form.get("note_type", "public")
    body = request.form.get("body", "").strip()

    if not body:
        flash("Empty note save नहीं होगा. Software भी कुछ standards रखता है.", "danger")
        return redirect(url_for("candidate_detail", candidate_code=candidate_code))

    # 1) Note save karo
    get_backend().insert("notes", {
        "candidate_id": candidate_code,
        "username": user["username"],
        "note_type": note_type,
        "body": body,
        "created_at": now_iso()
    })

    # 2) Public note ho to notifications bhejo, but fail ho to app crash mat karo
    candidate = get_candidate(candidate_code)
    if note_type == "public" and candidate:
        try:
            targets = [u for u in list_users() if u["role"] in {"manager", "tl"}]
            owner = find_user_by_recruiter_code(candidate.get("recruiter_code"))
            if owner:
                targets.append(owner)

            dedup = {}
            for t in targets:
                if t.get("user_id"):
                    dedup[t["user_id"]] = t

            preview = body[:90] + ("..." if len(body) > 90 else "")

            for i, target in enumerate(dedup.values(), start=1):
                get_backend().insert("notifications", {
                    "notification_id": f"N{int(datetime.now().timestamp()*1000)}{i}{random.randint(100,999)}",
                    "user_id": target["user_id"],
                    "title": f"Note updated: {candidate.get('full_name', candidate_code)}",
                    "message": f"{user.get('full_name', user['username'])} ({user.get('designation', '')}) added a public note on {candidate.get('full_name', candidate_code)}: {preview}",
                    "category": "note",
                    "status": "Unread",
                    "metadata": json.dumps({"candidate_id": candidate_code}),
                    "created_at": now_iso()
                })
        except Exception as e:
            print("Notification insert failed:", e)

    flash("Note saved successfully.", "success")
    return redirect(url_for("candidate_detail", candidate_code=candidate_code))
    user = current_user()
    note_type = request.form.get("note_type", "public")
    body = request.form.get("body", "").strip()
    if not body:
        flash("Empty note save नहीं होगा. Software भी कुछ standards रखता है.", "danger")
        return redirect(url_for("candidate_detail", candidate_code=candidate_code))
    get_backend().insert("notes", {
        "candidate_id": candidate_code,
        "username": user["username"],
        "note_type": note_type,
        "body": body,
        "created_at": now_iso()
    })
    candidate = get_candidate(candidate_code)
    if note_type == "public" and candidate:
        targets = [u for u in list_users() if u["role"] in {"manager", "tl"}]
        owner = find_user_by_recruiter_code(candidate.get("recruiter_code"))
        if owner:
            targets.append(owner)
        dedup = {}
        for t in targets:
            dedup[t["user_id"]] = t
        preview = body[:90] + ("..." if len(body) > 90 else "")
        for target in dedup.values():
            get_backend().insert("notifications", {
                "notification_id": f"N{int(datetime.now().timestamp()*1000)}{random.randint(10,99)}",
                "user_id": target["user_id"],
                "title": f"Note updated: {candidate['full_name']}",
                "message": f"{user['full_name']} ({user['designation']}) added a public note on {candidate['full_name']}: {preview}",
                "category": "note",
                "status": "Unread",
                "metadata": json.dumps({"candidate_id": candidate_code}),
                "created_at": now_iso()
            })
    flash("Note saved successfully.", "success")
    return redirect(url_for("candidate_detail", candidate_code=candidate_code))

@app.route("/candidate/create", methods=["POST"])
@login_required
def create_candidate():
    recruiter_code = request.form.get("recruiter_code", "").strip()
    owner = find_user_by_recruiter_code(recruiter_code) or current_user()
    all_rows = get_backend().list_rows("candidates")
    next_id = f"C{len(all_rows)+1:03d}"
    row = {
        "candidate_id": next_id,
        "full_name": request.form.get("full_name", "").strip(),
        "phone": request.form.get("phone", "").strip(),
        "qualification": request.form.get("qualification", "").strip(),
        "location": request.form.get("location", "").strip(),
        "experience": request.form.get("experience", "").strip(),
        "preferred_location": "",
        "qualification_level": "",
        "total_experience": "",
        "relevant_experience": "",
        "in_hand_salary": "",
        "career_gap": "",
        "documents_availability": "",
        "communication_skill": "",
        "relevant_experience_range": "",
        "relevant_in_hand_range": "",
        "submission_date": datetime.now().strftime("%Y-%m-%d"),
        "process": request.form.get("process", "").strip(),
        "recruiter_code": owner.get("recruiter_code", ""),
        "recruiter_name": owner.get("full_name", ""),
        "recruiter_designation": owner.get("designation", ""),
        "status": request.form.get("status", "New").strip() or "New",
        "all_details_sent": "Pending",
        "interview_reschedule_date": "",
        "is_duplicate": "0",
        "notes": request.form.get("notes", "").strip(),
        "resume_filename": "",
        "recording_filename": "",
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    if not row["full_name"] or not row["phone"]:
        flash("Candidate name and phone are required.", "danger")
        return redirect(url_for("candidates"))
    get_backend().insert("candidates", row)
    if owner:
        get_backend().insert("notifications", {
            "notification_id": f"N{int(datetime.now().timestamp()*1000)}{random.randint(10,99)}",
            "user_id": owner["user_id"],
            "title": "New candidate added",
            "message": f"{row['full_name']} was added to the CRM.",
            "category": "candidate",
            "status": "Unread",
            "metadata": json.dumps({"candidate_id": next_id}),
            "created_at": now_iso()
        })
    flash(f"Candidate {row['full_name']} added.", "success")
    return redirect(url_for("candidate_detail", candidate_code=next_id))

@app.route("/jds")
@login_required
def jds():
    status = request.args.get("status", "").strip()
    q = request.args.get("q", "").strip().lower()
    rows = [dict(r) for r in get_backend().list_rows("jd_master")]
    for row in rows:
        row["code"] = row.get("jd_id", "")
        row["title"] = row.get("job_title", "")
        row["status"] = "Open"
        row["experience_required"] = row.get("experience", "")
        row["payout"] = row.get("salary", "")
        row["payout_days"] = "60"
    if q:
        rows = [r for r in rows if q in " ".join([r.get("jd_id",""), r.get("job_title",""), r.get("company",""), r.get("location","")]).lower()]
    if status:
        rows = [r for r in rows if r.get("status") == status]
    status_choices = ["Open"]
    rows.sort(key=lambda x: x.get("code",""))
    return render_template("jds.html", jds=rows, status=status, status_choices=status_choices, q=request.args.get("q",""))

@app.route("/jd/create", methods=["POST"])
@login_required
def create_jd():
    rows = get_backend().list_rows("jd_master")
    jd_id = f"J{len(rows)+1:03d}"
    row = {
        "jd_id": jd_id,
        "job_title": request.form.get("job_title", "").strip(),
        "company": request.form.get("company", "").strip(),
        "location": request.form.get("location", "").strip(),
        "experience": request.form.get("experience", "").strip(),
        "salary": request.form.get("salary", "").strip(),
        "notes": request.form.get("notes", "").strip(),
        "created_at": now_iso()
    }
    if not row["job_title"] or not row["company"]:
        flash("JD title and company are required.", "danger")
        return redirect(url_for("jds"))
    get_backend().insert("jd_master", row)
    flash(f"JD {row['job_title']} added.", "success")
    return redirect(url_for("jds"))

@app.route("/interviews")
@login_required
def interviews():
    current_stage = request.args.get("stage", "All")
    interviews = []
    candidates_by_id = candidate_map()
    jd_by_id = {j.get("jd_id"): j for j in get_backend().list_rows("jd_master")}
    for row in get_backend().list_rows("interviews"):
        item = dict(row)
        candidate = candidates_by_id.get(item.get("candidate_id")) or {}
        jd = jd_by_id.get(item.get("jd_id")) or {}
        item["full_name"] = candidate.get("full_name", "")
        item["title"] = jd.get("job_title", candidate.get("process", ""))
        if current_stage != "All" and item.get("stage") != current_stage:
            continue
        interviews.append(item)
    interviews.sort(key=lambda x: x.get("scheduled_at",""))
    return render_template("interviews.html", interviews=interviews, current_stage=current_stage)

@app.route("/interview/create", methods=["POST"])
@login_required
def create_interview():
    candidate_id = request.form.get("candidate_id", "").strip()
    jd_id = request.form.get("jd_id", "").strip()
    stage = request.form.get("stage", "").strip() or "Screening"
    scheduled_at = request.form.get("scheduled_at", "").strip()
    if not candidate_id or not scheduled_at:
        flash("Candidate ID and schedule time are required.", "danger")
        return redirect(url_for("interviews"))
    rows = get_backend().list_rows("interviews")
    row = {
        "interview_id": f"I{len(rows)+1:03d}",
        "candidate_id": candidate_id,
        "jd_id": jd_id,
        "stage": stage,
        "scheduled_at": scheduled_at,
        "status": "Scheduled",
        "created_at": now_iso()
    }
    get_backend().insert("interviews", row)
    flash(f"Interview scheduled for {candidate_id}.", "success")
    return redirect(url_for("interviews"))

@app.route("/tasks")
@login_required
def tasks():
    user = current_user()
    rows = []
    users_by_id = user_map("user_id")
    raw = get_backend().list_rows("tasks")
    for t in raw:
        item = dict(t)
        assigned_user = users_by_id.get(item.get("assigned_to_user_id")) or {}
        item["full_name"] = assigned_user.get("full_name", item.get("assigned_to_name", ""))
        item["due_at"] = item.get("due_date", "")
        if user["role"] != "manager" and item.get("assigned_to_user_id") != user["user_id"]:
            continue
        rows.append(item)
    rows.sort(key=lambda x: x.get("due_at",""))
    return render_template("tasks.html", tasks=rows)

@app.route("/task/create", methods=["POST"])
@login_required
def create_task():
    target = get_user(request.form.get("assigned_to_username", "").strip())
    creator = current_user()
    if not target:
        flash("Assigned username not found.", "danger")
        return redirect(url_for("tasks"))
    rows = get_backend().list_rows("tasks")
    row = {
        "task_id": f"T{len(rows)+1:03d}",
        "title": request.form.get("title", "").strip(),
        "description": request.form.get("description", "").strip(),
        "assigned_to_user_id": target["user_id"],
        "assigned_to_name": target["full_name"],
        "assigned_by_user_id": creator["user_id"],
        "assigned_by_name": creator["full_name"],
        "status": request.form.get("status", "Open").strip() or "Open",
        "priority": request.form.get("priority", "Normal").strip() or "Normal",
        "due_date": request.form.get("due_date", datetime.now().strftime("%Y-%m-%d")),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    if not row["title"]:
        flash("Task title required.", "danger")
        return redirect(url_for("tasks"))
    get_backend().insert("tasks", row)
    get_backend().insert("notifications", {
        "notification_id": f"N{int(datetime.now().timestamp()*1000)}{random.randint(10,99)}",
        "user_id": target["user_id"],
        "title": "Task assigned",
        "message": row["title"],
        "category": "task",
        "status": "Unread",
        "metadata": json.dumps({"task_id": row["task_id"]}),
        "created_at": now_iso()
    })
    flash("Task added.", "success")
    return redirect(url_for("tasks"))

@app.route("/submissions")
@login_required
def submissions():
    candidates_by_id = candidate_map()
    jds_by_id = {j.get("jd_id"): j for j in get_backend().list_rows("jd_master")}
    rows = []
    for s in get_backend().list_rows("submissions"):
        item = dict(s)
        candidate = candidates_by_id.get(item.get("candidate_id")) or {}
        jd = jds_by_id.get(item.get("jd_id")) or {}
        item["full_name"] = candidate.get("full_name", "")
        item["title"] = jd.get("job_title", candidate.get("process", ""))
        rows.append(item)
    rows.sort(key=lambda x: x.get("submitted_at", ""), reverse=True)
    return render_template("submissions.html", submissions=rows)

@app.route("/notifications")
@login_required
def notifications_page():
    user = current_user()
    rows = user_notifications(user)
    return render_template("notifications.html", notifications=rows)

@app.route("/notifications/mark-all-read")
@login_required
def mark_all_read():
    user = current_user()
    for n in user_notifications(user):
        get_backend().update_where("notifications", {"notification_id": n["notification_id"]}, {"status": "Read"})
    flash("All notifications marked as read.", "success")
    return redirect(url_for("notifications_page"))

@app.route("/chat", methods=["GET", "POST"])
@login_required
def chat_page():
    user = current_user()
    users = [u for u in list_users() if u["username"] != user["username"]]
    users.sort(key=lambda x: (x["role"], x.get("full_name","")))
    selected = request.args.get("with") or (users[0]["username"] if users else None)
    if request.method == "POST":
        recipient = request.form.get("recipient")
        body = request.form.get("body", "").strip()
        if recipient and body:
            get_backend().insert("messages", {
                "sender_username": user["username"],
                "recipient_username": recipient,
                "body": body,
                "created_at": now_iso()
            })
            flash("Message sent.", "success")
            return redirect(url_for("chat_page", **{"with": recipient}))
    convo = []
    if selected:
        for m in get_backend().list_rows("messages"):
            if {m.get("sender_username"), m.get("recipient_username")} == {user["username"], selected}:
                item = dict(m)
                item["sender_name"] = (get_user(item.get("sender_username")) or {}).get("full_name", item.get("sender_username"))
                item["recipient_name"] = (get_user(item.get("recipient_username")) or {}).get("full_name", item.get("recipient_username"))
                convo.append(item)
    convo.sort(key=lambda x: x.get("created_at", ""))
    return render_template("chat.html", users=users, selected=selected, convo=convo)

@app.route("/admin")
@login_required
@manager_required
def admin_page():
    users = sorted(list_users(), key=lambda x: (x["role"], x.get("full_name","")))
    notes = get_backend().list_rows("notes")
    counts = []
    for user in users:
        public_count = len([n for n in notes if n.get("username") == user["username"] and n.get("note_type") == "public"])
        private_count = len([n for n in notes if n.get("username") == user["username"] and n.get("note_type") == "private"])
        counts.append({"full_name": user["full_name"], "public_count": public_count, "private_count": private_count})
    return render_template("admin.html", users=users, notes_count=counts)

@app.route("/admin/impersonate", methods=["POST"])
@login_required
@manager_required
def impersonate_login():
    username = request.form.get("username", "").strip()
    target = get_user(username)
    manager = get_user(session.get("username"))
    if not target or not manager:
        flash("Target account not found.", "danger")
        return redirect(url_for("admin_page"))
    session["impersonator"] = manager["username"]
    session["impersonated_as"] = target["username"]
    flash(f"Now viewing as {target['full_name']}.", "success")
    return redirect(url_for("dashboard"))

@app.route("/admin/stop-impersonation")
@login_required
def stop_impersonation():
    if session.get("impersonator"):
        original = session.get("impersonator")
        session.pop("impersonated_as", None)
        session.pop("impersonator", None)
        session["username"] = original
        flash("Returned to manager account.", "success")
    return redirect(url_for("admin_page"))

@app.route("/module/<slug>")
@login_required
def module_page(slug):
    module = MODULE_SUMMARIES.get(slug)
    if not module:
        abort(404)
    dialer_candidates = enrich_candidates() if slug == "dialer" else []
    meeting_feed = []
    if slug == 'meeting-room':
        meeting_feed = [
            {"name": "Ritika joined", "time": "03:00 PM", "state": "Joined"},
            {"name": "Mohit joined", "time": "03:02 PM", "state": "Joined"},
            {"name": "Barnali left", "time": "03:11 PM", "state": "Left"},
            {"name": "Neha joined", "time": "03:14 PM", "state": "Joined"},
        ]
    return render_template("module_page.html", module=module, slug=slug, dialer_candidates=dialer_candidates, meeting_feed=meeting_feed)

@app.route("/blueprint")
@login_required
def blueprint_page():
    blueprint_path = BASE_DIR / "docs" / "MEGA_BLUEPRINT_120_PLUS_FEATURES.md"
    context_path = BASE_DIR / "docs" / "CROSS_CHAT_MASTER_CONTEXT.txt"
    blueprint_text = blueprint_path.read_text(encoding="utf-8")
    context_text = context_path.read_text(encoding="utf-8")
    return render_template("blueprint.html", blueprint_text=blueprint_text, context_text=context_text)

@app.route("/preview")
def preview_page():
    demo_users = sorted(list_users(), key=lambda x: (x["role"], x.get("full_name","")))
    return render_template("preview.html", demo_users=demo_users)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
